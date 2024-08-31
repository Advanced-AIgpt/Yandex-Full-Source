#!/usr/bin/env python
# encoding: utf-8
import json
from openpyxl import Workbook
from openpyxl.styles import Font
from random import sample
from nirvana.job_context import context


# Применяется в https://nirvana.yandex-team.ru/flow/d74e9b3d-ee70-4856-8f3b-e1445b8698ff
# ...на разметке, полученной через https://nirvana.yandex-team.ru/flow/2e6becd0-027c-4e14-9734-02d1850ea94f

def extract(references, assignments, threshold, ref_shift=0, validation_size=200):  # TODO: validation_size
    stat = {'total': 0,    # Всего референсов
            'correct': 0,  # Тасков, корректно размеченных Толокерами
            'missing_ref': 0,  # Тасков, для которых референс оказался непроставленным
            'correct_over_thr': 0,  # Тасков, размеченных корректно с уверенностью превышающей threshold
            'incorrect_over_thr': 0,  # Тасков, размеченных корректно с уверенностью превышающей threshold
            'accuracy': None,  # Процент тасков размеченных корректно
            }

    honeypots = []
    hints = make_hints_worksheet()
    validation = make_validation_worksheet()
    validation_ids = set(sample(xrange(ref_shift, ref_shift+len(references)), validation_size))

    for task in assignments:
        kind, id = task['inputValues']['key'].split('_')
        if kind == 'ref':
            try:
                idx = int(id) - ref_shift
                ref_res = references[idx]['result']
            except KeyError:
                #print 'id:', id, 'ref:', references[int(id)]
                stat['missing_ref'] += 1
                continue
            else:
                if ref_res is None:
                    stat['missing_ref'] += 1
                    continue
                if ref_res == 'awesome':  # available only on 'good' answers
                    ref_res = 'good'
                stat['total'] += 1

            task_res = task['outputValues']['result']
            if ref_res == task_res:
                stat['correct'] += 1
                if task['probability']['result'] >= threshold:
                    stat['correct_over_thr'] += 1
                    honeypots.append(make_honeypot(task, ref_res))
            else:
                if ref_res not in ('good', 'neutral', 'bad'):
                    print 'ref_res:', ref_res
                if task['probability']['result'] >= threshold:
                    stat['incorrect_over_thr'] += 1
                    hints.append(make_hint(task, ref_res))

            if idx in validation_ids:
                validation.append(make_validation(task, ref_res))

    stat['accuracy'] = float(stat['correct']) / stat['total']
    return honeypots, hints, validation, stat


def make_honeypot(task, result):
    return {"inputValues": task['inputValues'],
            "knownSolutions": [
                {"outputValues": {"result": result}}
            ]
    }


def make_hints_worksheet():
    ws = Workbook().active  # active worksheet
    ws.append(['key', 'context_2', 'context_1', 'context_0', 'reply', 'toloka_result', 'reference_result'])
    header_font = Font(bold=True)
    for row in ws['A1:G1']:
        for cell in row:
            cell.font = header_font

    for col in xrange(1, 5):
        name = chr(ord('A') + col)
        ws.column_dimensions[name].width = 30

    return ws


def make_hint(task, result):
    inp = task['inputValues']
    return [inp['key'], inp['context_2'], inp['context_1'], inp['context_0'], inp['reply'],
            task['outputValues']['result'], result]


def make_validation_worksheet():
    ws = Workbook().active  # active worksheet
    ws.append(['key', 'context_2', 'context_1', 'context_0', 'reply', 'match', 'toloka_result', 'reference_result'])
    header_font = Font(bold=True)
    for row in ws['A1:H1']:
        for cell in row:
            cell.font = header_font

    for col in xrange(1, 5):
        name = chr(ord('A') + col)
        ws.column_dimensions[name].width = 30

    return ws


def make_validation(task, result):
    inp = task['inputValues']
    match = task['probability']['result']
    if task['outputValues']['result'] != result:
        match = -match
    return [inp['key'], inp['context_2'], inp['context_1'], inp['context_0'], inp['reply'],
            match, task['outputValues']['result'], result]


def main():
    ctx = context()
    inputs = ctx.get_inputs()
    outputs = ctx.get_outputs()
    params = ctx.get_parameters()

    references = json.load(open(inputs.get('references')))
    assignments = json.load(open(inputs.get('assignments')))
    threshold = params['threshold']
    ref_shift = params['ref_shift']
    validation_size = params['validation_size']

    honeypots, hints, validation, stat = extract(references, assignments, threshold, ref_shift, validation_size)

    json.dump(honeypots, open(outputs.get('honeypots.json'), 'w'), sort_keys=True, indent=2) #, ensure_ascii=False, encoding='utf-8')

    hints.parent.save(outputs.get('hints.xlsx'))

    validation.parent.save(outputs.get('validation.xlsx'))

    json.dump(stat, open(outputs.get('stat'), 'w'), sort_keys=True, indent=2)


if __name__ == '__main__':
    main()

